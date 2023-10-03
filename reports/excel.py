import datetime
import itertools
import operator

from django.db.models import Count
from django.utils import timezone
from openpyxl.workbook import Workbook

from robots.models import Robot


def create_xlsx_report(name: str):
    week_robots_sold = (Robot.objects.filter(created__gte=timezone.now() - datetime.timedelta(weeks=1))
                        .values('model', 'version', )
                        .annotate(count=Count('created'))
                        .order_by('model'))
    wb = Workbook()
    # delete default sheet
    for defaults in wb.sheetnames:
        wb.remove(wb[defaults])

    # Unfortunately, django doesn't have something like QuerySet.groupby('model')
    for model, model_group in itertools.groupby(week_robots_sold, key=operator.itemgetter('model')):  # group by model
        ws = wb.create_sheet(model)
        ws.append(['Модель', 'Версия', 'Количество за неделю'])
        for robot in model_group:
            ws.append(list(robot.values()))

    wb.save(name)
    return wb.mime_type
